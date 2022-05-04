import pandas as pd
import os
from datetime import datetime
from gensim.test.utils import get_tmpfile
from openpyxl.styles import  PatternFill, Font
from gensim.models.fasttext import FastText 
from gensim.models.fasttext import load_facebook_model
from openpyxl import load_workbook
from openpyxl import Workbook
from utils.constants import EXCELFILE
from utils.constants import SHEETNAME1
from utils.constants import SHEETNAME2
from utils.constants import SHEETNAME3
from utils.constants import ROOT_DIR
from utils.constants import MODEL_DIRS

def check_curr_best_models_save_and_delete(model,modelName,modelPath):
    
    #check current best models
    best_model_dict = copy_eval_results_and_check_best_models()
    
    #save the model if it is the current best one
    if modelName in best_model_dict.keys():
        if model is not None:
            save_model(model,modelName,modelPath)
        
    #delete not best models to save space
    
    delete_not_best_models(best_model_dict, MODEL_DIRS)
    
def copy_excel_sort_by_modelname(source,target):
    
    df1=pd.read_excel(source,SHEETNAME1)
    df2=pd.read_excel(source,SHEETNAME2)
    df3=pd.read_excel(source,SHEETNAME3)
    
    
    df1_new=sort_by_model_name(df1)
    df2_new=sort_by_model_name(df2)
    df3_new=sort_by_model_name(df3)
    
    append_eval_results_to_excel(target,SHEETNAME1,df1_new)
    
    append_eval_results_to_excel(target,SHEETNAME3,df3_new)
    
    append_eval_results_to_excel(target,SHEETNAME2,df2_new)
    
    wb = load_workbook(target)
    for ws in wb.worksheets:
        if ws.title not in [SHEETNAME1,SHEETNAME2,SHEETNAME3]:
            wb.remove(ws)
    wb.save(target)

def sort_by_model_name(df):
    pts=[x for x in df['modelName'] if x.startswith('pt')]
    fts=[x for x in df['modelName'] if x.startswith('ft')]
    Cfts=[x for x in df['modelName'] if x.startswith('Cft')]
    berts=[x for x in df['modelName'] if x.startswith('bert')]
    Cberts=[x for x in df['modelName'] if x.startswith('Cbert')]
    
    
    pt=(df.loc[df['modelName'].isin(pts)]).sort_values(by=['modelName'])
    ft=(df.loc[df['modelName'].isin(fts)]).sort_values(by=['modelName'])
    Cft=(df.loc[df['modelName'].isin(Cfts)]).sort_values(by=['modelName'])
    bert=(df.loc[df['modelName'].isin(berts)]).sort_values(by=['modelName'])
    Cbert=(df.loc[df['modelName'].isin(Cberts)]).sort_values(by=['modelName'])
    
    df_new=pd.concat([pt,bert,Cbert,ft,Cft])
    return df_new

def color_the_best_metric(excelfile,sheetname,columnname,maxmetric):
    wb = load_workbook(excelfile)
    ws = wb[sheetname]
   
    ## Create a dictionary of column names
    ColNames = {}
    Current  = 0
    for COL in ws.iter_cols(1, ws.max_column):
        ColNames[COL[0].value] = Current
        Current += 1

    ## Now you can access by column name
    ## (My data has a column named 'Dogs')
    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if (row_cells[ColNames[columnname]].value==maxmetric):
            #print(row_cells[ColNames[columnname]])
            row_cells[ColNames[columnname]].fill = PatternFill("solid", fgColor="DDDDDD")
            row_cells[ColNames[columnname]].font = Font(b=True)
            #print(row_cells[ColNames[columnname]].fill)
            wb.save(excelfile)
            
def copy_excel(source,target):
    wb = load_workbook(source)
    wb.save(target)
    
def copy_eval_results_and_check_best_models():
    #copy the evaluation results
    source=ROOT_DIR+'/'+EXCELFILE
    now = datetime.now()
    dt_string = now.strftime("%Y-%m-%d-%H-%M")
    target=source[:-5]+'_Copy'+dt_string +'.xlsx'
    copy_excel_sort_by_modelname(source,target)
    
    #check best models and color the best metrics
    best_models=[]
    metrics1=['precision@k=1','recall@k=1','f1@k=1','precision@k=3','recall@k=3','f1@k=3','precision@k=6','recall@k=6','f1@k=6',]
    metrics2=['cv_f1','precision','recall','f1']
    
    df1=pd.read_excel(target,SHEETNAME1)
    df2=pd.read_excel(target,SHEETNAME3)
    
    for m in metrics1:
        idxmax=df1[m].idxmax()
        maxmetric=df1[m].max()
        modelname = df1['modelName'][idxmax]
        best_models.append(modelname)
        color_the_best_metric(target,SHEETNAME1,m,maxmetric)
        
    for m in metrics2:
        idxmax=df2[m].idxmax()
        modelname = df2['modelName'][idxmax]
        maxmetric=df2[m].max()
        best_models.append(modelname)
        color_the_best_metric(target,SHEETNAME3,m,maxmetric)
        
    best_model_dict=pd.DataFrame(best_models, columns=["name"]).groupby('name').size().to_dict()
    
    print('Current best models are: ',best_model_dict)
    return best_model_dict

def check_best_models():
    best_models=[]
    metrics1=['precision@k=1','recall@k=1','f1@k=1','precision@k=3','recall@k=3','f1@k=3','precision@k=6','recall@k=6','f1@k=6',]
    metrics2=['cv_f1','precision','recall','f1']
    df1=pd.read_excel(ROOT_DIR+'/'+EXCELFILE,SHEETNAME1)
    df2=pd.read_excel(ROOT_DIR+'/'+EXCELFILE,SHEETNAME3)
    for m in metrics1:
        idxmax=df1[m].idxmax()
        modelname = df1['modelName'][idxmax]
        best_models.append(modelname)
    for m in metrics2:
        idxmax=df2[m].idxmax()
        modelname = df2['modelName'][idxmax]
        best_models.append(modelname)
    best_model_dict=pd.DataFrame(best_models, columns=["name"]).groupby('name').size().to_dict()
    
    print('Current best models are: ',best_model_dict)
    return best_model_dict

def delete_not_best_models(best_model_dict, DIRS):
    best_models = best_model_dict.keys()
    for a_dir in DIRS:
        subdirs = [name for name in os.listdir(a_dir)
            if os.path.isdir(os.path.join(a_dir, name))]
        for subdir in subdirs:
            if subdir not in best_models:
                files=os.listdir(a_dir+'/'+subdir)
                for file in files:
                    if not file.endswith('.csv'):
                        os.remove(a_dir+'/'+subdir+'/'+file)
                        print('MODEL '+subdir+' removed!')   
    

def save_model(model,modelName,modelPath):
      
    create_dir(modelPath)
    
    fname = get_tmpfile(modelPath)
    print("Saving model... ",modelPath)
    #save fasttext model
    model.save(fname)
    #save as word2vec format
    model.wv.save_word2vec_format(modelPath[:-5]+'vec',binary=False)
    
    print("Model saved: "+modelName)
def create_dir(file_path):
    
    directory = os.path.dirname(file_path)
    
    if not os.path.exists(directory):
        print(directory+" not existed, creating...")
        os.makedirs(directory)

def load_bin_model(modelPath):
    print("Loading the model...")
    #path=os.getcwd()+modelPath
    model =  load_facebook_model(modelPath)
    #model = FastText.load_fasttext_format(path,full_model=True)
    print("Loading done.")
    print_model_info(modelPath,model)
    
    return model

def load_ft_model(modelPath):
    print("Loading the model...")
    #path=os.getcwd()+modelPath
    model =  FastText.load(modelPath)
    print("Loading done.")
    print_model_info(modelPath,model)
    return model

def print_model_info(modelPath,model):
    print("====================================================================")
    print("Model Path: ", modelPath)
    print("window: ",model.window)
    print("embed size: ",len(model.wv["mann"]))
    print("sg: "  , model.sg)
    print("epochs: ", model.epochs)
    print("min_count: ",model.vocabulary.min_count)
    print("negative sampling: ", model.negative)
    print("min_n:",model.wv.min_n)
    print("max_n:",model.wv.max_n)
    print("Vocabulary size: ",len(model.wv.vectors_vocab))
    print("====================================================================")

def generate_domain_vocab(synspairs,rootdir):
    list_vocab=[]
    for file in synspairs:
        file=rootdir+'/'+file
        r = open(file,'r',encoding='utf-8')
        lines = r.readlines()
        for line in lines:
            if not line.startswith(": "):
                words=line.split(" ###### ")
                if (len(words)==2):
                    list_vocab.append(words[0].strip())
                    list_vocab.append(words[1].strip())
                else:
                    raise Exception("False: len(words)==2",words)
    #remove duplicates   
    list_vocab = list(dict.fromkeys(list_vocab))
    #print("The domain vocabulary is of size ",len(list_vocab))
    return list_vocab

def generate_syns_dictionary(vocab,synspairs,rootdir):
    syns_dict={}
    for entity in vocab:
        syns=[]
        for file in synspairs:
            file=rootdir+'/'+file
            r = open(file,'r',encoding='utf-8')
            lines = r.readlines()
            for line in lines:
                if not line.startswith(": "):
                    words=line.split(" ###### ")
                    w0=words[0].strip()
                    w1=words[1].strip()
                    if w0==entity:
                        syns.append(w1)
                    if w1==entity:
                        syns.append(w0)
        #remove duplicates   
        syns= list(dict.fromkeys(syns))
        syns_dict.update({entity:syns})
    return syns_dict



def compute_oov(vocab,model,modelName):
    oovs=[]
    for entity in vocab:
        oovs.append(1-int(entity in model.wv.vocab))
    oov=round(sum(oovs)/len(oovs),4)*100
    #print(str(oov)+"% of the domain vocabulary are out-of-vacabulary of the model "+modelName)
    return oov

def append_eval_results_to_excel(excelfile,sheetname,df):
    print('Saving the evaluation results...')
    
    if not os.path.isfile(excelfile):
        print('The excel file is not existing, creating a new excel file...'+excelfile)       
        wb = Workbook()
        wb.save(excelfile)
        
        
    wb = load_workbook(excelfile)
    if not (sheetname in wb.sheetnames):
        print('The worksheet is not existing, creating a new worksheet...'+sheetname) 
        ws1 = wb.create_sheet(sheetname)
        ws1.title = sheetname
        wb.save(excelfile)
    
            
    book = load_workbook(excelfile)
    writer = pd.ExcelWriter(excelfile, engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    startRow=writer.sheets[sheetname].max_row
    
    if (startRow==1):
        df.to_excel(writer,sheet_name=sheetname, startrow=startRow-1, index = False,header= True)
        
    else:
        df.to_excel(writer,sheet_name=sheetname, startrow=startRow, index = False,header= False)
        

    writer.save()
                

